import json
import os
import sys
from openpyxl import load_workbook


def main():
    here = os.path.dirname(__file__)
    file_path = os.path.abspath(os.path.join(here, '..', 'assets', 'Site_Split.xlsx'))

    if not os.path.exists(file_path):
        print(json.dumps({"error": "File not found", "path": file_path}))
        sys.exit(1)

    wb = load_workbook(filename=file_path, data_only=True, read_only=True)
    try:
        summary = {"file": os.path.basename(file_path), "sheets": []}

        for name in wb.sheetnames:
            ws = wb[name]
            rows_preview = []
            headers = None

            for idx, row in enumerate(ws.iter_rows(values_only=True)):
                if headers is None:
                    headers = [str(c) if c is not None else None for c in row]
                    continue

                record = {}
                for i, h in enumerate(headers or []):
                    key = h if h not in (None, "", "None") else f"col_{i+1}"
                    val = row[i] if i < len(row) else None
                    record[key] = val
                rows_preview.append(record)

                if len(rows_preview) >= 10:
                    break

            total_rows = max(ws.max_row - 1, 0) if ws.max_row else 0
            summary["sheets"].append({
                "name": name,
                "rows": total_rows,
                "preview": rows_preview
            })

        print(json.dumps(summary, indent=2, default=str))
    finally:
        wb.close()


if __name__ == "__main__":
    main()
