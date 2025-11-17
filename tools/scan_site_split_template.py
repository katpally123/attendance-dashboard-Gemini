import sys
import json
import re
from typing import Optional, Dict
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

LABELS = [
    "Regular HC (Cohort Expected)",
    "Regular HC Present (Excluding Swaps)",
    "Shift Swap Out ",
    "Shift Swap Expected",
    "Shift Swap Present ",
    "VTO",
    "VET Accepted",
    "VET Present",
    "MET Expected",
    "MET Present",
]

DEPTS = ["Inbound","DA","ICQA","CRETs"]

def _norm(s: Optional[str]) -> str:
    if s is None:
        return ''
    s = str(s)
    s = s.replace('\u00A0',' ').replace('\xa0',' ')
    s = re.sub(r'[\-/]+',' ', s)
    s = re.sub(r'\s+',' ', s).strip().upper()
    s = s.replace('CRÉTS','CRETS')
    return s

def find_anchor_row(ws) -> int:
    for r in range(1, min(ws.max_row, 50)+1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and _norm(v) == 'ATTENDANCE DETAILS':
            return r
    # fallback to 3 (as in legacy)
    return 3

def build_column_map(ws, anchor_row: int) -> Dict[tuple, int]:
    targets: Dict[tuple, int] = {}
    for c in range(2, ws.max_column+1):
        tokens = set()
        for rr in (anchor_row, anchor_row+1, anchor_row+2):
            v = ws.cell(row=rr, column=c).value
            if isinstance(v, str):
                for t in _norm(v).split():
                    tokens.add(t)
        if not tokens:
            continue
        if 'TOTAL' in tokens or 'PERCENT' in tokens or 'PCT' in tokens:
            continue
        def set_once(key, *need):
            if key in targets:
                return
            if all(n in tokens for n in need):
                targets[key] = c

        set_once(("Inbound","AMZN"), "INBOUND","AMZN")
        set_once(("Inbound","TEMP"), "INBOUND","TEMP")
        set_once(("DA","AMZN"), "DA","AMZN")
        set_once(("DA","TEMP"), "DA","TEMP")
        set_once(("ICQA","AMZN"), "ICQA","AMZN")
        set_once(("ICQA","TEMP"), "ICQA","TEMP")
        # CRETs / IXD group
        if ("CRETS","AMZN") not in targets and {"CRETS","AMZN"}.issubset(tokens):
            targets[("CRETS","AMZN")] = c
        if ("CRETS","TEMP") not in targets and {"CRETS","TEMP"}.issubset(tokens):
            targets[("CRETS","TEMP")] = c
        if ("CRETS","AMZN") not in targets and {"IXD","AMZN"}.issubset(tokens):
            targets[("CRETS","AMZN")] = c
        if ("CRETS","TEMP") not in targets and {"IXD","TEMP"}.issubset(tokens):
            targets[("CRETS","TEMP")] = c
    return targets

def find_label_row(ws, label: str) -> Optional[int]:
    want = _norm(label)
    for r in range(1, ws.max_row+1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and _norm(v) == want:
            return r
    return None

def main(path: str):
    wb = load_workbook(path, data_only=False)
    # Prefer sheet with 'Attendance' in title if present
    ws = wb.active
    for name in wb.sheetnames:
        if 'attendance' in name.lower():
            ws = wb[name]
            break

    anchor = find_anchor_row(ws)
    col_targets = build_column_map(ws, anchor)
    col_letters = {k: get_column_letter(v) for k, v in col_targets.items()}

    mapping = {}
    for label in LABELS:
        rr = find_label_row(ws, label)
        if not rr:
            continue
        per_dept = {}
        for d in DEPTS:
            am = col_letters.get((d,'AMZN'))
            tp = col_letters.get((d,'TEMP'))
            if am and tp:
                per_dept[d] = { 'AMZN': f"{am}{rr}", 'TEMP': f"{tp}{rr}" }
        if per_dept:
            mapping[label] = per_dept

    print("# ---- SITE_SPLIT_MAP (paste into site_split_export.py) ----")
    print("SITE_SPLIT_MAP = ")
    print(json.dumps(mapping, indent=2, sort_keys=False))

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python tools/scan_site_split_template.py <path-to-xlsx>")
        sys.exit(1)
    main(sys.argv[1])
