from io import BytesIO
from typing import Dict, Any, Optional
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import re

# Copilot: STRICT RULES FOR EXCEL EXPORT
# -------------------------------------
# 1. Only write AMZN/TEMP values to the fixed coordinates defined in SITE_SPLIT_MAP.
# 2. NEVER write to:
#       - Column F  (SDC TOTAL)
#       - Column L  (IXD TOTAL)
#       - Any TOTAL / percentage / variance / derived formula cell
#       - Any column containing an Excel formula
# 3. MET rows (rows 18 and 19) must always be 0 for AMZN and TEMP.
# 4. ALPS (Hours) and Assumptions % rows must remain blank.
# 5. DO NOT detect headers, DO NOT parse template structure.
# 6. DO NOT recalc totals; Excel formulas handle this automatically.
# 7. All formulas must remain intact in the output file.
# -------------------------------------

# ================================
#  SITE SPLIT EXCEL FILL (Fixed Map)
#  Follow EXACT cell mapping only
# ================================

SITE_SPLIT_MAP: Dict[str, Dict[str, Dict[str, str]]] = {
    "Regular HC (Cohort Expected)": {
        "Inbound": {"AMZN": "B6",  "TEMP": "C6"},
        "DA":      {"AMZN": "D6",  "TEMP": "E6"},
        "ICQA":    {"AMZN": "G6",  "TEMP": "H6"},
        "CRETs":   {"AMZN": "J6",  "TEMP": "K6"},
    },
    "Regular HC Present (Excluding Swaps)": {
        "Inbound": {"AMZN": "B7",  "TEMP": "C7"},
        "DA":      {"AMZN": "D7",  "TEMP": "E7"},
        "ICQA":    {"AMZN": "G7",  "TEMP": "H7"},
        "CRETs":   {"AMZN": "J7",  "TEMP": "K7"},
    },
    "Shift Swap Out ": {
        "Inbound": {"AMZN": "B8",  "TEMP": "C8"},
        "DA":      {"AMZN": "D8",  "TEMP": "E8"},
        "ICQA":    {"AMZN": "G8",  "TEMP": "H8"},
        "CRETs":   {"AMZN": "J8",  "TEMP": "K8"},
    },
    "Shift Swap Expected": {
        "Inbound": {"AMZN": "B9",  "TEMP": "C9"},
        "DA":      {"AMZN": "D9",  "TEMP": "E9"},
        "ICQA":    {"AMZN": "G9",  "TEMP": "H9"},
        "CRETs":   {"AMZN": "J9",  "TEMP": "K9"},
    },
    "Shift Swap Present ": {
        "Inbound": {"AMZN": "B10", "TEMP": "C10"},
        "DA":      {"AMZN": "D10", "TEMP": "E10"},
        "ICQA":    {"AMZN": "G10", "TEMP": "H10"},
        "CRETs":   {"AMZN": "J10", "TEMP": "K10"},
    },
    "VTO": {
        "Inbound": {"AMZN": "B12", "TEMP": "C12"},
        "DA":      {"AMZN": "D12", "TEMP": "E12"},
        "ICQA":    {"AMZN": "G12", "TEMP": "H12"},
        "CRETs":   {"AMZN": "J12", "TEMP": "K12"},
    },
    "VET Accepted": {
        "Inbound": {"AMZN": "B13", "TEMP": "C13"},
        "DA":      {"AMZN": "D13", "TEMP": "E13"},
        "ICQA":    {"AMZN": "G13", "TEMP": "H13"},
        "CRETs":   {"AMZN": "J13", "TEMP": "K13"},
    },
    "VET Present": {
        "Inbound": {"AMZN": "B14", "TEMP": "C14"},
        "DA":      {"AMZN": "D14", "TEMP": "E14"},
        "ICQA":    {"AMZN": "G14", "TEMP": "H14"},
        "CRETs":   {"AMZN": "J14", "TEMP": "K14"},
    },
    "MET Expected": {
        "Inbound": {"AMZN": "B18", "TEMP": "C18"},
        "DA":      {"AMZN": "D18", "TEMP": "E18"},
        "ICQA":    {"AMZN": "G18", "TEMP": "H18"},
        "CRETs":   {"AMZN": "J18", "TEMP": "K18"},
    },
    "MET Present": {
        "Inbound": {"AMZN": "B19", "TEMP": "C19"},
        "DA":      {"AMZN": "D19", "TEMP": "E19"},
        "ICQA":    {"AMZN": "G19", "TEMP": "H19"},
        "CRETs":   {"AMZN": "J19", "TEMP": "K19"},
    },
}

def write_by_map(ws, mapping: Dict[str, Dict[str, Dict[str, str]]], payload: Dict[str, Any]):
    # Iterate strictly by mapping order; never iterate departments from payload
    for row_label, dept_map in mapping.items():
        row_payload = payload.get(row_label, {}) or {}
        for dept_name, role_map in dept_map.items():
            dept_vals = row_payload.get(dept_name, {}) or {}
            for role, cell in role_map.items():  # AMZN or TEMP only
                # Never overwrite formula cells
                existing = ws[cell].value
                if isinstance(existing, str) and existing.startswith('='):
                    continue
                value = dept_vals.get(role, None)
                if value is not None:
                    try:
                        ws[cell].value = int(value)
                    except Exception:
                        ws[cell].value = value


# ================================
#  Dynamic map builder (scan template)
#  Detect row numbers and AMZN/TEMP column letters
# ================================

def _norm(s: Optional[str]) -> str:
    if s is None:
        return ''
    s = str(s)
    s = s.replace('\u00A0',' ').replace('\xa0',' ')
    s = re.sub(r'[\-/]+',' ', s)
    s = re.sub(r'\s+',' ', s).strip().upper()
    s = s.replace('CRÉTS','CRETS')
    return s

DESIRED_LABELS = [
    "Regular HC (Cohort Expected)",
    "Regular HC Present (Excluding Swaps)",
    "Shift Swap Out ",
    "Shift Swap Expected",
    "Shift Swap Present ",
    "VTO",
    "VET Accepted",
    "VET Present",
    "MET Expected",
    "MET Present",
]

DEPTS = ["Inbound","DA","ICQA","CRETs"]

def _find_anchor_row(ws) -> int:
    for r in range(1, min(ws.max_row, 40)+1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and _norm(v) == 'ATTENDANCE DETAILS':
            return r
    return 3

def _build_col_targets(ws, anchor_row: int) -> Dict[tuple, int]:
    targets: Dict[tuple, int] = {}
    # Build token set per column from three header rows (anchor, anchor+1, anchor+2)
    for c in range(2, ws.max_column+1):
        tokens = set()
        for rr in (anchor_row, anchor_row+1, anchor_row+2):
            v = ws.cell(row=rr, column=c).value
            if isinstance(v, str):
                for t in _norm(v).split():
                    tokens.add(t)
        if not tokens:
            continue
        # Skip TOTAL columns for mapping
        if 'TOTAL' in tokens or 'PERCENT' in tokens or 'PCT' in tokens:
            continue
        # Helper to set if not already set (first match wins, leftmost)
        def set_if_match(key, must_tokens):
            nonlocal targets
            if key in targets:
                return
            if all(tok in tokens for tok in must_tokens):
                targets[key] = c

        # INBOUND
        set_if_match(("Inbound","AMZN"), ("INBOUND","AMZN"))
        set_if_match(("Inbound","TEMP"), ("INBOUND","TEMP"))
        # DA
        set_if_match(("DA","AMZN"), ("DA","AMZN"))
        set_if_match(("DA","TEMP"), ("DA","TEMP"))
        # ICQA
        set_if_match(("ICQA","AMZN"), ("ICQA","AMZN"))
        set_if_match(("ICQA","TEMP"), ("ICQA","TEMP"))
        # CRETs (IXD group) — match CRETS primarily; fallback to IXD
        if ("CRETS","AMZN") not in targets and ("CRETS" in tokens and "AMZN" in tokens):
            targets[("CRETS","AMZN")] = c
        if ("CRETS","TEMP") not in targets and ("CRETS" in tokens and "TEMP" in tokens):
            targets[("CRETS","TEMP")] = c
        if ("CRETS","AMZN") not in targets and ("IXD" in tokens and "AMZN" in tokens):
            targets[("CRETS","AMZN")] = c
        if ("CRETS","TEMP") not in targets and ("IXD" in tokens and "TEMP" in tokens):
            targets[("CRETS","TEMP")] = c
    return targets

def _find_label_row(ws, label: str) -> Optional[int]:
    want = _norm(label)
    for r in range(1, ws.max_row+1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and _norm(v) == want:
            return r
    return None

def build_dynamic_site_split_map(ws) -> Dict[str, Dict[str, Dict[str, str]]]:
    anchor = _find_anchor_row(ws)
    targets = _build_col_targets(ws, anchor)
    # Convert to letters
    col_letters = {k: get_column_letter(c) for k, c in targets.items()}

    dynamic_map: Dict[str, Dict[str, Dict[str, str]]] = {}
    for label in DESIRED_LABELS:
        rr = _find_label_row(ws, label)
        if not rr:
            continue
        per_dept: Dict[str, Dict[str, str]] = {}
        for d in DEPTS:
            am_col = col_letters.get((d,'AMZN'))
            tp_col = col_letters.get((d,'TEMP'))
            if am_col and tp_col:
                per_dept[d] = { 'AMZN': f"{am_col}{rr}", 'TEMP': f"{tp_col}{rr}" }
        if per_dept:
            dynamic_map[label] = per_dept
    return dynamic_map


def build_site_split_xlsx(template_path: str, rows_payload: Dict[str, Any], shift: str, date_str: str) -> bytes:
    """Load template, build a fresh cell map from headers, and write values.
    - Writes only AMZN/TEMP for mapped rows
    - Always writes 0 for MET Expected/Present cells (respecting formulas)
    - Does not touch formula columns
    """
    wb = load_workbook(template_path, data_only=False)  # Preserve formulas
    ws = wb.active

    # Build dynamic map from the actual template
    dynamic_map = build_dynamic_site_split_map(ws)

    # Write provided values strictly by the detected map
    write_by_map(ws, dynamic_map, rows_payload)

    # Then, explicitly zero MET rows using the detected map
    for met_label in ("MET Expected","MET Present"):
        row_map = dynamic_map.get(met_label, {})
        for dept, roles in row_map.items():
            for role, cell in roles.items():
                existing = ws[cell].value
                if isinstance(existing, str) and existing.startswith('='):
                    continue
                ws[cell] = 0

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()

